import os
import io
import boto3
import openpyxl


s3_bucket = "my-bucket"
filename = "excelfile.xlsx"


def lambda_handler(event, context):
    s3 = boto3.client('s3')
    #files_response = s3.list_objects_v2(Bucket="my-bucket")
    filething = s3.get_object(Bucket=s3_bucket, Key=filename)
    filedata = filething['Body'].read()
    workbook = openpyxl.load_workbook(io.BytesIO(filedata), data_only=True)
    worksheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
    
    return {"filething": str(filething),
            "tabs": workbook.sheetnames,
            "A1": worksheet["A1"].value}
